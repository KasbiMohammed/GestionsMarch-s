"""
Export Excel avec OpenPyXL
Fonctions pour générer des fichiers Excel pour les marchés et étapes
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List
from datetime import datetime
from app.models.market import Market


def create_styled_header(sheet, headers, row_num=1):
    """
    Crée un en-tête stylisé pour une feuille Excel
    
    Args:
        sheet: Feuille Excel
        headers: Liste des en-têtes
        row_num: Numéro de ligne pour l'en-tête
    """
    # Style pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajuster la largeur des colonnes
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width


def export_markets_to_excel(markets: List[Market], filepath: str):
    """
    Exporte une liste de marchés vers un fichier Excel
    
    Args:
        markets: Liste des marchés à exporter
        filepath: Chemin du fichier de sortie
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Marchés"
    
    # En-têtes
    headers = [
        "Numéro",
        "Objet",
        "Maître d'ouvrage",
        "Type",
        "Mode de passation",
        "Montant estimé",
        "Montant définitif",
        "Budget",
        "Service responsable",
        "Statut",
        "Progression (%)",
        "Date de publication",
        "Date d'ouverture",
        "Date d'attribution",
        "Date de début",
        "Date de fin prévue",
        "Date de fin réelle",
        "Observations"
    ]
    
    create_styled_header(ws, headers)
    
    # Données
    row_num = 2
    for market in markets:
        ws.cell(row=row_num, column=1, value=market.market_number)
        ws.cell(row=row_num, column=2, value=market.object)
        ws.cell(row=row_num, column=3, value=market.master_of_work)
        ws.cell(row=row_num, column=4, value=market.market_type.value if market.market_type else "")
        ws.cell(row=row_num, column=5, value=market.procurement_method.value if market.procurement_method else "")
        ws.cell(row=row_num, column=6, value=market.estimated_amount)
        ws.cell(row=row_num, column=7, value=market.definitive_amount or "")
        ws.cell(row=row_num, column=8, value=market.budget or "")
        ws.cell(row=row_num, column=9, value=market.responsible_service or "")
        ws.cell(row=row_num, column=10, value=market.status.value if market.status else "")
        ws.cell(row=row_num, column=11, value=market.progress_percentage)
        ws.cell(row=row_num, column=12, value=market.publication_date.strftime("%d/%m/%Y") if market.publication_date else "")
        ws.cell(row=row_num, column=13, value=market.opening_date.strftime("%d/%m/%Y") if market.opening_date else "")
        ws.cell(row=row_num, column=14, value=market.attribution_date.strftime("%d/%m/%Y") if market.attribution_date else "")
        ws.cell(row=row_num, column=15, value=market.start_date.strftime("%d/%m/%Y") if market.start_date else "")
        ws.cell(row=row_num, column=16, value=market.expected_end_date.strftime("%d/%m/%Y") if market.expected_end_date else "")
        ws.cell(row=row_num, column=17, value=market.actual_end_date.strftime("%d/%m/%Y") if market.actual_end_date else "")
        ws.cell(row=row_num, column=18, value=market.observations or "")
        
        row_num += 1
    
    # Sauvegarder le fichier
    wb.save(filepath)


def _format_date(dt) -> str:
    """Formate une date pour l'export."""
    return dt.strftime("%d/%m/%Y") if dt else ""


def export_plannings_to_excel(plannings: List, filepath: str):
    """
    Exporte une liste de planifications vers un fichier Excel

    Args:
        plannings: Liste des planifications à exporter
        filepath: Chemin du fichier de sortie
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Planifications"

    headers = [
        "Numéro",
        "Exercice",
        "Intitulé",
        "Type projet",
        "Type procédure",
        "Budget estimatif",
        "Source financement",
        "Service demandeur",
        "Responsable",
        "Priorité",
        "Statut",
        "Lancement",
        "Ouverture plis",
        "Attribution",
        "Notification",
        "Ordre de service",
        "Début",
        "Fin",
        "Observations",
    ]

    create_styled_header(ws, headers)

    row_num = 2
    for p in plannings:
        ws.cell(row=row_num, column=1, value=p.planning_number)
        ws.cell(row=row_num, column=2, value=p.fiscal_year)
        ws.cell(row=row_num, column=3, value=p.title)
        ws.cell(row=row_num, column=4, value=p.project_type.value if p.project_type else "")
        ws.cell(row=row_num, column=5, value=p.procedure_type.value if p.procedure_type else "")
        ws.cell(row=row_num, column=6, value=p.estimated_budget)
        ws.cell(row=row_num, column=7, value=p.funding_source or "")
        ws.cell(row=row_num, column=8, value=p.requesting_service_name or "")
        ws.cell(row=row_num, column=9, value=p.responsible_name or "")
        ws.cell(row=row_num, column=10, value=p.priority.value if p.priority else "")
        ws.cell(row=row_num, column=11, value=p.status.value if p.status else "")
        ws.cell(row=row_num, column=12, value=_format_date(p.launch_date))
        ws.cell(row=row_num, column=13, value=_format_date(p.bid_opening_date))
        ws.cell(row=row_num, column=14, value=_format_date(p.attribution_date))
        ws.cell(row=row_num, column=15, value=_format_date(p.notification_date))
        ws.cell(row=row_num, column=16, value=_format_date(p.service_order_date))
        ws.cell(row=row_num, column=17, value=_format_date(p.start_date))
        ws.cell(row=row_num, column=18, value=_format_date(p.end_date))
        ws.cell(row=row_num, column=19, value=p.observations or "")
        row_num += 1

    wb.save(filepath)


def export_stages_to_excel(stages: List, filepath: str):
    """
    Exporte une liste d'étapes vers un fichier Excel
    
    Args:
        stages: Liste des étapes à exporter
        filepath: Chemin du fichier de sortie
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Étapes"
    
    # En-têtes
    headers = [
        "ID Marché",
        "Nom de l'étape",
        "Code",
        "Catégorie",
        "Statut",
        "Progression (%)",
        "Date prévue",
        "Date réelle",
        "Retard (jours)",
        "Niveau d'alerte",
        "Responsable",
        "Validé",
        "Observations"
    ]
    
    create_styled_header(ws, headers)
    
    # Données
    row_num = 2
    for stage in stages:
        ws.cell(row=row_num, column=1, value=stage.market_id)
        ws.cell(row=row_num, column=2, value=stage.name)
        ws.cell(row=row_num, column=3, value=stage.code or "")
        ws.cell(row=row_num, column=4, value=stage.category.value if stage.category else "")
        ws.cell(row=row_num, column=5, value=stage.status.value if stage.status else "")
        ws.cell(row=row_num, column=6, value=stage.progress_percentage)
        ws.cell(row=row_num, column=7, value=stage.planned_date.strftime("%d/%m/%Y") if stage.planned_date else "")
        ws.cell(row=row_num, column=8, value=stage.actual_date.strftime("%d/%m/%Y") if stage.actual_date else "")
        ws.cell(row=row_num, column=9, value=stage.delay_days)
        ws.cell(row=row_num, column=10, value=stage.alert_level or "")
        ws.cell(row=row_num, column=11, value=stage.responsible_id or "")
        ws.cell(row=row_num, column=12, value="Oui" if stage.is_validated else "Non")
        ws.cell(row=row_num, column=13, value=stage.observations or "")
        
        row_num += 1
    
    # Sauvegarder le fichier
    wb.save(filepath)


def export_companies_to_excel(companies: List, filepath: str):
    """
    Exporte une liste d'entreprises vers un fichier Excel
    
    Args:
        companies: Liste des entreprises à exporter
        filepath: Chemin du fichier de sortie
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Entreprises"
    
    # En-têtes
    headers = [
        "ID Marché",
        "Nom de l'entreprise",
        "RC",
        "IF",
        "Adresse",
        "Téléphone",
        "Email",
        "Montant de l'offre",
        "Rang",
        "Attributaire",
        "Offre anormalement basse",
        "Offre anormalement haute",
        "Score technique",
        "Score financier",
        "Score total",
        "Observations"
    ]
    
    create_styled_header(ws, headers)
    
    # Données
    row_num = 2
    for company in companies:
        ws.cell(row=row_num, column=1, value=company.market_id)
        ws.cell(row=row_num, column=2, value=company.name)
        ws.cell(row=row_num, column=3, value=company.rc_number or "")
        ws.cell(row=row_num, column=4, value=company.if_number or "")
        ws.cell(row=row_num, column=5, value=company.address or "")
        ws.cell(row=row_num, column=6, value=company.phone or "")
        ws.cell(row=row_num, column=7, value=company.email or "")
        ws.cell(row=row_num, column=8, value=company.offer_amount or "")
        ws.cell(row=row_num, column=9, value=company.offer_rank or "")
        ws.cell(row=row_num, column=10, value="Oui" if company.is_attributed else "Non")
        ws.cell(row=row_num, column=11, value="Oui" if company.is_abnormally_low else "Non")
        ws.cell(row=row_num, column=12, value="Oui" if company.is_abnormally_high else "Non")
        ws.cell(row=row_num, column=13, value=company.technical_score or "")
        ws.cell(row=row_num, column=14, value=company.financial_score or "")
        ws.cell(row=row_num, column=15, value=company.total_score or "")
        ws.cell(row=row_num, column=16, value=company.observations or "")
        
        row_num += 1
    
    # Sauvegarder le fichier
    wb.save(filepath)


def export_dashboard_to_excel(stats: dict, filepath: str):
    """
    Exporte les statistiques du dashboard vers un fichier Excel
    
    Args:
        stats: Dictionnaire des statistiques du dashboard
        filepath: Chemin du fichier de sortie
    """
    wb = Workbook()
    
    # Feuille Statistiques Marchés
    ws1 = wb.active
    ws1.title = "Statistiques Marchés"
    
    headers1 = ["Métrique", "Valeur"]
    create_styled_header(ws1, headers1)
    
    row_num = 2
    market_stats = stats.get('markets', {})
    ws1.cell(row=row_num, column=1, value="Total Marchés")
    ws1.cell(row=row_num, column=2, value=market_stats.get('total', 0))
    row_num += 1
    ws1.cell(row=row_num, column=1, value="En Cours")
    ws1.cell(row=row_num, column=2, value=market_stats.get('en_cours', 0))
    row_num += 1
    ws1.cell(row=row_num, column=1, value="Terminés")
    ws1.cell(row=row_num, column=2, value=market_stats.get('termine', 0))
    row_num += 1
    ws1.cell(row=row_num, column=1, value="En Retard")
    ws1.cell(row=row_num, column=2, value=market_stats.get('en_retard', 0))
    row_num += 1
    ws1.cell(row=row_num, column=1, value="En Attente")
    ws1.cell(row=row_num, column=2, value=market_stats.get('en_attente', 0))
    row_num += 1
    ws1.cell(row=row_num, column=1, value="Annulés")
    ws1.cell(row=row_num, column=2, value=market_stats.get('annule', 0))
    row_num += 1
    ws1.cell(row=row_num, column=1, value="Suspendus")
    ws1.cell(row=row_num, column=2, value=market_stats.get('suspendu', 0))
    
    # Feuille Montants
    ws2 = wb.create_sheet("Montants")
    create_styled_header(ws2, headers1)
    
    row_num = 2
    amounts = stats.get('amounts', {})
    ws2.cell(row=row_num, column=1, value="Montant Estimé Total")
    ws2.cell(row=row_num, column=2, value=amounts.get('total_estimated', 0))
    row_num += 1
    ws2.cell(row=row_num, column=1, value="Montant Définitif Total")
    ws2.cell(row=row_num, column=2, value=amounts.get('total_definitive', 0))
    row_num += 1
    ws2.cell(row=row_num, column=1, value="Montant Attribué")
    ws2.cell(row=row_num, column=2, value=amounts.get('total_attributed', 0))
    row_num += 1
    ws2.cell(row=row_num, column=1, value="Montant Payé")
    ws2.cell(row=row_num, column=2, value=amounts.get('total_paid', 0))
    row_num += 1
    ws2.cell(row=row_num, column=1, value="Budget Restant")
    ws2.cell(row=row_num, column=2, value=amounts.get('remaining', 0))
    row_num += 1
    ws2.cell(row=row_num, column=1, value="Écart Estimation/Attribution")
    ws2.cell(row=row_num, column=2, value=amounts.get('variance', 0))
    
    # Feuille Planification
    ws3 = wb.create_sheet("Planification")
    create_styled_header(ws3, headers1)
    
    row_num = 2
    planning = stats.get('planning', {})
    ws3.cell(row=row_num, column=1, value="Total Planifications")
    ws3.cell(row=row_num, column=2, value=planning.get('total', 0))
    row_num += 1
    ws3.cell(row=row_num, column=1, value="Budget Planifié")
    ws3.cell(row=row_num, column=2, value=planning.get('budget', 0))
    row_num += 1
    ws3.cell(row=row_num, column=1, value="Validées")
    ws3.cell(row=row_num, column=2, value=planning.get('validated', 0))
    row_num += 1
    ws3.cell(row=row_num, column=1, value="Programmées")
    ws3.cell(row=row_num, column=2, value=planning.get('programmed', 0))
    
    # Sauvegarder le fichier
    wb.save(filepath)
